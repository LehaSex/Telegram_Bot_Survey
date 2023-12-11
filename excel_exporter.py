import openpyxl
import db
import json

async def export_excel():
    # init excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    ws1 = wb.create_sheet("Вопросы")
    ws2 = wb.create_sheet("Ответы")

    # users
    ws["A1"] = "ID пользователя"
    ws["B1"] = "Имя"
    ws["C1"] = "Фамилия"
    ws["D1"] = "Никнейм в боте"
    ws["E1"] = "Дата регистрации"
    ws["F1"] = "Администратор"

    # questions
    ws1["A1"] = "ID вопроса"
    ws1["B1"] = "Текст вопроса"
    ws1["C1"] = "Варианты ответов"

    # answers
    ws2["A1"] = "ID ответа"
    ws2["B1"] = "ID пользователя"
    ws2["C1"] = "ID вопроса"
    ws2["D1"] = "Ответ"

    # users
    db_cursor = db.db_connection.cursor()
    db_cursor.execute("SELECT * FROM users")
    users = db_cursor.fetchall()
    db_cursor.close()
    for i in range(len(users)):
        if users[i][4] == None:
            ws["A" + str(i + 2)] = users[i][1]
        else:
           ws["A" + str(i + 2)] = users[i][4] 
        ws["B" + str(i + 2)] = users[i][2]
        ws["C" + str(i + 2)] = users[i][3]
        ws["D" + str(i + 2)] = users[i][7]
        ws["E" + str(i + 2)] = users[i][5]
        ws["F" + str(i + 2)] = "Да" if users[i][6] == 1 or 2 else "Нет"

    # questions
    db_cursor = db.db_connection.cursor()
    db_cursor.execute("SELECT * FROM questions")
    questions = db_cursor.fetchall()
    db_cursor.close()
    for i in range(len(questions)):
        ws1["A" + str(i + 2)] = questions[i][0]
        ws1["B" + str(i + 2)] = questions[i][1]
        # answers
        answers = json.loads(questions[i][2])
        answers_str = ""
        for j in range(len(answers)):
            answers_str += answers[j] + ", "
        ws1["C" + str(i + 2)] = answers_str 

    # answers
    db_cursor = db.db_connection.cursor()
    db_cursor.execute("SELECT * FROM answers")
    answers = db_cursor.fetchall()
    db_cursor.close()
    for i in range(len(answers)):
        # get answer from answers
        ws2["A" + str(i + 2)] = answers[i][0]
        # find user, that chat_id == answers[i][1]
        user = None
        for j in range(len(users)):
            if users[j][1] == answers[i][1]:
                user = users[j]
                break
        # link to user in users sheet
        if user[4] == None:
            ws2["B" + str(i + 2)] = '=HYPERLINK("{}", "{}")'.format("#'Пользователи'!A" + str(j + 2), user[1])
        else:
            ws2["B" + str(i + 2)] = '=HYPERLINK("{}", "{}")'.format("#'Пользователи'!A" + str(j + 2), "@" + user[4])
        # find question, that id == answers[i][2]
        question = None
        for j in range(len(questions)):
            if questions[j][0] == answers[i][2]:
                question = questions[j]
                break
        # link to question in questions sheet
        ws2["C" + str(i + 2)] = '=HYPERLINK("{}", "{}")'.format("#'Вопросы'!A" + str(j + 2), question[1][:40] + "...")
        ws2["D" + str(i + 2)] = answers[i][3]


    # pretty print
    await excel_pretty_print(ws, 1.2)
    await excel_pretty_print(ws1, 0.3)
    await excel_pretty_print(ws2, 1.2)
    # save excel
    wb.save("export.xlsx")

async def excel_pretty_print(ws, value):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * value
        ws.column_dimensions[column].width = adjusted_width
